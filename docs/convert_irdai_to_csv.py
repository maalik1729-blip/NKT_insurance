"""
IRDAI Death Claims XLSX → CSV Converter
Converts downloaded IRDAI XLSX files to same format as existing dataset.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import openpyxl
import numpy as np

os.makedirs("docs/fetched", exist_ok=True)

DOWNLOADED = {
    "2021-22": "docs/fetched/death_claims_2021-22.xlsx",
    "2020-21": "docs/fetched/death_claims_2020-21.xlsx",
    "2019-20": "docs/fetched/death_claims_2019-20.xlsx",
    "2018-19": "docs/fetched/death_claims_2018-19.xlsx",
    "2022-23": "docs/fetched/death_claims_2022-23.xlsx",
}

def parse_sheet(wb, sheet_name, year, category):
    """Parse Indl-DC or Group DC sheet into rows."""
    ws = wb[sheet_name]
    rows = []
    current_insurer = None
    
    # Collect all rows first
    all_rows = list(ws.iter_rows(values_only=True))
    
    # Find header row (row with 'Life Insurer')
    header_row_idx = None
    for i, row in enumerate(all_rows):
        if any(str(c) == 'Life Insurer' for c in row if c):
            header_row_idx = i
            break
    
    if header_row_idx is None:
        print(f"  Could not find header in {sheet_name}")
        return []
    
    # Data starts after header rows (usually 3-4 header rows)
    data_start = header_row_idx + 3
    
    for row in all_rows[data_start:]:
        if row is None or all(c is None for c in row):
            continue
        
        # Col A = insurer name (or empty if continuation)
        # Col B = year
        insurer_cell = str(row[0]).strip() if row[0] is not None else ''
        year_cell    = str(row[1]).strip() if row[1] is not None else ''
        
        # Skip rows that are totals/headers
        if any(x in insurer_cell.lower() for x in ['industry', 'total', 'pvt', 'private', 'public']):
            # Still capture industry totals
            pass
        
        if insurer_cell and insurer_cell != 'None':
            current_insurer = insurer_cell
        
        if not year_cell or year_cell == 'None' or not current_insurer:
            continue
        
        # Only process rows where year matches expected pattern
        if '-' not in year_cell and len(year_cell) < 4:
            continue
        
        try:
            # Extract numeric values (handle None/empty)
            def safe_float(val):
                if val is None or str(val).strip() in ['', 'None', '-', 'N/A']:
                    return 0.0
                try:
                    return float(val)
                except:
                    return 0.0
            
            def safe_int(val):
                if val is None or str(val).strip() in ['', 'None', '-', 'N/A']:
                    return 0
                try:
                    return int(float(val))
                except:
                    return 0
            
            # Column mapping based on IRDAI format:
            # A=Insurer, B=Year, C=Pend_Start_No, D=Pend_Start_Amt,
            # E=Intimated_No, F=Intimated_Amt, G=Total_No, H=Total_Amt,
            # I=Paid_No, J=Paid_Amt, K=Repud_No, L=Repud_Amt,
            # M=Rejected_No, N=Rejected_Amt, O=Unclaimed_No, P=Unclaimed_Amt,
            # Q=Pend_End_No, R=Pend_End_Amt, S=CSR_No%, T=CSR_Amt%
            # (columns may vary slightly between years)
            
            cols = list(row)
            while len(cols) < 24:
                cols.append(None)
            
            pend_start_no  = safe_int(cols[2])
            pend_start_amt = safe_float(cols[3])
            intim_no       = safe_int(cols[4])
            intim_amt      = safe_float(cols[5])
            total_no       = safe_int(cols[6])
            total_amt      = safe_float(cols[7])
            paid_no        = safe_int(cols[8])
            paid_amt       = safe_float(cols[9])
            repud_no       = safe_int(cols[10])
            repud_amt      = safe_float(cols[11])
            reject_no      = safe_int(cols[12])
            reject_amt     = safe_float(cols[13])
            unclaim_no     = safe_int(cols[14])
            unclaim_amt    = safe_float(cols[15])
            pend_end_no    = safe_int(cols[16])
            pend_end_amt   = safe_float(cols[17])
            
            # Calculate ratios
            csr_no  = paid_no  / total_no  if total_no  > 0 else 0.0
            csr_amt = paid_amt / total_amt if total_amt > 0 else 0.0
            rep_rej_ratio_no  = (repud_no + reject_no)  / total_no  if total_no  > 0 else 0.0
            rep_rej_ratio_amt = (repud_amt + reject_amt) / total_amt if total_amt > 0 else 0.0
            pend_ratio_no  = pend_end_no  / total_no  if total_no  > 0 else 0.0
            pend_ratio_amt = pend_end_amt / total_amt if total_amt > 0 else 0.0
            
            rows.append({
                'life_insurer': current_insurer,
                'year': year_cell,
                'claims_pending_start_no': pend_start_no,
                'claims_pending_start_amt': pend_start_amt,
                'claims_intimated_no': intim_no,
                'claims_intimated_amt': intim_amt,
                'total_claims_no': total_no,
                'total_claims_amt': total_amt,
                'claims_paid_no': paid_no,
                'claims_paid_amt': paid_amt,
                'claims_repudiated_no': repud_no,
                'claims_repudiated_amt': repud_amt,
                'claims_rejected_no': reject_no,
                'claims_rejected_amt': reject_amt,
                'claims_unclaimed_no': unclaim_no,
                'claims_unclaimed_amt': unclaim_amt,
                'claims_pending_end_no': pend_end_no,
                'claims_pending_end_amt': pend_end_amt,
                'claims_paid_ratio_no': round(csr_no, 16),
                'claims_paid_ratio_amt': round(csr_amt, 16),
                'claims_repudiated_rejected_ratio_no': round(rep_rej_ratio_no, 16),
                'claims_repudiated_rejected_ratio_amt': round(rep_rej_ratio_amt, 16),
                'claims_pending_ratio_no': round(pend_ratio_no, 16),
                'claims_pending_ratio_amt': round(pend_ratio_amt, 16),
                'category': category,
            })
        except Exception as e:
            print(f"  Row error: {e} | row={[str(c)[:10] for c in row[:8]]}")
    
    return rows

# =========================================================
# Process all downloaded files
# =========================================================
all_individual = []
all_group = []

for year, filepath in DOWNLOADED.items():
    if not os.path.exists(filepath):
        print(f"Missing: {filepath}")
        continue
    
    print(f"\nProcessing {year}...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    
    # Individual Death Claims
    if 'Indl-DC' in wb.sheetnames:
        rows = parse_sheet(wb, 'Indl-DC', year, 'Individual Death Claims')
        print(f"  Individual DC rows: {len(rows)}")
        all_individual.extend(rows)
    
    # Group Death Claims
    if 'Group DC' in wb.sheetnames:
        rows = parse_sheet(wb, 'Group DC', year, 'Group Death Claims')
        print(f"  Group DC rows: {len(rows)}")
        all_group.extend(rows)

# =========================================================
# Save CSVs
# =========================================================
print("\n" + "=" * 60)
print("Saving CSVs...")
print("=" * 60)

if all_individual:
    df_ind = pd.DataFrame(all_individual)
    df_ind = df_ind.drop_duplicates(subset=['life_insurer', 'year', 'category'])
    df_ind = df_ind[df_ind['total_claims_no'] > 0]
    fname = "docs/fetched/irdai_individual_death_claims.csv"
    df_ind.to_csv(fname, index=False)
    print(f"Individual Death Claims: {len(df_ind)} rows -> {fname}")
    print(df_ind[df_ind['life_insurer'].str.contains('LIC', case=False, na=False)][['life_insurer','year','claims_paid_ratio_no','claims_intimated_no','claims_paid_no']].to_string())

if all_group:
    df_grp = pd.DataFrame(all_group)
    df_grp = df_grp.drop_duplicates(subset=['life_insurer', 'year', 'category'])
    df_grp = df_grp[df_grp['total_claims_no'] > 0]
    fname = "docs/fetched/irdai_group_death_claims.csv"
    df_grp.to_csv(fname, index=False)
    print(f"\nGroup Death Claims: {len(df_grp)} rows -> {fname}")
    print(df_grp[df_grp['life_insurer'].str.contains('LIC', case=False, na=False)][['life_insurer','year','claims_paid_ratio_no','claims_intimated_no','claims_paid_no']].to_string())

print("\nDone! Check docs/fetched/ folder for CSV files.")
