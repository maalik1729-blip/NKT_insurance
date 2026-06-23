"""
Fix & Merge IRDAI 2022-23 data with existing archive dataset.
2022-23 XLSX has data for 2021-22 AND 2022-23 rows — extract 2022-23 only.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import openpyxl

def parse_death_claims_sheet(wb, sheet_name, target_year, category):
    """Extract only target_year rows from a death claims sheet."""
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    rows = []
    current_insurer = None
    
    for row in all_rows:
        if row is None or all(c is None for c in row):
            continue
        
        col_a = str(row[0]).strip() if row[0] is not None else ''
        col_b = str(row[1]).strip() if row[1] is not None else ''
        
        # Skip pure header rows
        if col_a in ['Life Insurer', 'INDIVIDUAL DEATH CLAIMS', 'GROUP DEATH CLAIMS',
                     'Indl-DC', 'Group DC', '', 'None']:
            if col_a not in ['', 'None']:
                pass
        
        # Update current insurer if col A has a name
        if col_a and col_a not in ['None', '', 'Life Insurer'] and not col_a.startswith('INDIVIDUAL') and not col_a.startswith('GROUP'):
            # Check it's not a number
            try:
                float(col_a)
            except:
                current_insurer = col_a
        
        # Only process rows matching target year
        if col_b != target_year or not current_insurer:
            continue
        
        def safe_float(v):
            if v is None or str(v).strip() in ['', 'None', '-']: return 0.0
            try: return float(v)
            except: return 0.0
        
        def safe_int(v):
            if v is None or str(v).strip() in ['', 'None', '-']: return 0
            try: return int(float(v))
            except: return 0
        
        cols = list(row)
        while len(cols) < 24:
            cols.append(None)
        
        pend_start_no  = safe_int(cols[2])
        pend_start_amt = safe_float(cols[3])
        intim_no       = safe_int(cols[4])
        intim_amt      = safe_float(cols[5])
        total_no       = safe_int(cols[6])
        total_amt      = safe_float(cols[7])
        paid_no        = safe_int(cols[8])
        paid_amt       = safe_float(cols[9])
        repud_no       = safe_int(cols[10])
        repud_amt      = safe_float(cols[11])
        reject_no      = safe_int(cols[12])
        reject_amt     = safe_float(cols[13])
        unclaim_no     = safe_int(cols[14])
        unclaim_amt    = safe_float(cols[15])
        pend_end_no    = safe_int(cols[16])
        pend_end_amt   = safe_float(cols[17])
        
        if total_no == 0:
            continue
        
        csr_no  = paid_no  / total_no  if total_no  > 0 else 0.0
        csr_amt = paid_amt / total_amt if total_amt > 0 else 0.0
        rrr_no  = (repud_no  + reject_no)  / total_no  if total_no  > 0 else 0.0
        rrr_amt = (repud_amt + reject_amt) / total_amt if total_amt > 0 else 0.0
        prn     = pend_end_no  / total_no  if total_no  > 0 else 0.0
        pra     = pend_end_amt / total_amt if total_amt > 0 else 0.0
        
        rows.append({
            'life_insurer': current_insurer,
            'year': target_year,
            'claims_pending_start_no':   pend_start_no,
            'claims_pending_start_amt':  pend_start_amt,
            'claims_intimated_no':       intim_no,
            'claims_intimated_amt':      intim_amt,
            'total_claims_no':           total_no,
            'total_claims_amt':          total_amt,
            'claims_paid_no':            paid_no,
            'claims_paid_amt':           paid_amt,
            'claims_repudiated_no':      repud_no,
            'claims_repudiated_amt':     repud_amt,
            'claims_rejected_no':        reject_no,
            'claims_rejected_amt':       reject_amt,
            'claims_unclaimed_no':       unclaim_no,
            'claims_unclaimed_amt':      unclaim_amt,
            'claims_pending_end_no':     pend_end_no,
            'claims_pending_end_amt':    pend_end_amt,
            'claims_paid_ratio_no':      round(csr_no, 16),
            'claims_paid_ratio_amt':     round(csr_amt, 16),
            'claims_repudiated_rejected_ratio_no':  round(rrr_no, 16),
            'claims_repudiated_rejected_ratio_amt': round(rrr_amt, 16),
            'claims_pending_ratio_no':   round(prn, 16),
            'claims_pending_ratio_amt':  round(pra, 16),
            'category': category,
        })
    
    return rows

# =====================================================
# Extract 2022-23 from the downloaded file
# =====================================================
print("Extracting 2022-23 data from IRDAI XLSX...")
wb = openpyxl.load_workbook("docs/fetched/death_claims_2022-23.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}")

# Debug: print raw rows from Indl-DC sheet to see what years are there
ws = wb['Indl-DC']
all_rows = list(ws.iter_rows(values_only=True))
print("\nSample rows from Indl-DC sheet:")
for i, row in enumerate(all_rows):
    if i < 70 and row[1] is not None:
        print(f"  Row {i}: insurer={str(row[0])[:20]}, year={row[1]}, total_no={row[6]}, paid_no={row[8]}")

print("\nUnique years in Indl-DC:")
years = set()
for row in all_rows:
    if row[1] is not None and str(row[1]).strip() not in ['', 'None', 'Year']:
        years.add(str(row[1]).strip())
print(f"  {sorted(years)}")
