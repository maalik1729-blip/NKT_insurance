"""
IRDAI Death Claims Finder v4
- Uses known folder IDs to find claims data
- Downloads and checks XLSX content
"""
import sys, io, requests, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from bs4 import BeautifulSoup

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
os.makedirs("docs/fetched", exist_ok=True)

BASE = "https://irdai.gov.in"
GROUP_ID = "37343"

def get(url, timeout=15):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        return r
    except Exception as e:
        print(f"  ERR: {e}")
        return None

def find_xlsx_links(html):
    soup = BeautifulSoup(html, 'html.parser')
    links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        txt  = a.get_text(strip=True)[:80]
        if any(x in href.lower() for x in ['.xlsx', '.xls', '.zip', '.csv']):
            full = href if href.startswith('http') else BASE + href
            links.append({'text': txt, 'url': full})
    return links

# =========================================================
# Step 1: Try IRDAI folder IDs for claims data
# Known folder IDs from premium pages: 365644, 367479
# Try systematic nearby IDs
# =========================================================
print("=" * 60)
print("STEP 1: Searching IRDAI document folders for claims")
print("=" * 60)

# Try known folder IDs and nearby ones
known_folders = [365644, 367479, 388048]
claim_folders_to_try = []
for base_id in known_folders:
    for offset in range(-20, 21, 2):
        claim_folders_to_try.append(base_id + offset)
# Also try some specific IDs
claim_folders_to_try += [370000, 371000, 372000, 373000, 374000, 375000, 
                          376000, 377000, 378000, 379000, 380000, 381000,
                          382000, 383000, 384000, 385000, 386000, 387000,
                          360000, 361000, 362000, 363000, 364000]

# Deduplicate
claim_folders_to_try = sorted(set(claim_folders_to_try))

claim_links = []
for folder_id in claim_folders_to_try:
    url = f"{BASE}/documents/{GROUP_ID}/{folder_id}"
    r = get(url, timeout=8)
    if r and r.status_code == 200 and len(r.content) > 10000:
        links = find_xlsx_links(r.text)
        death_links = [l for l in links if any(k in (l['text']+l['url']).lower() 
                       for k in ['claim','death','individual','group','settlement','mortality'])]
        if death_links:
            print(f"\nFOLDER {folder_id} has CLAIM links!")
            for l in death_links[:5]:
                print(f"  [{l['text'][:60]}]")
                print(f"   -> {l['url'][:110]}")
                claim_links.append(l)
        elif links:
            # Show first link to understand what's in this folder
            print(f"Folder {folder_id}: {links[0]['text'][:50]} (+ {len(links)-1} more xlsx)")

# =========================================================
# Step 2: Try the IRDAI "Life3" page which has 757KB content
# Extract all links from it
# =========================================================
print("\n" + "=" * 60)
print("STEP 2: Deep scan of IRDAI life3 page")
print("=" * 60)

r = get("https://irdai.gov.in/en/life3")
if r and r.status_code == 200:
    soup = BeautifulSoup(r.text, 'html.parser')
    # Find ALL links - not just downloads
    all_links = [(a.get_text(strip=True)[:70], 
                  a['href'] if a['href'].startswith('http') else BASE+a['href']) 
                 for a in soup.find_all('a', href=True) if a.get_text(strip=True)]
    
    print(f"Total links on life3 page: {len(all_links)}")
    # Filter for claims-related
    claim_related = [(t,u) for t,u in all_links if any(k in (t+u).lower() 
                     for k in ['claim','death','settlement','individual','group','mortality'])]
    print(f"Claims-related links: {len(claim_related)}")
    for t,u in claim_related[:15]:
        print(f"  [{t}] -> {u[:100]}")
    
    # Also look for any iframes or data attributes
    iframes = soup.find_all('iframe')
    print(f"\nIframes found: {len(iframes)}")
    for iframe in iframes[:3]:
        print(f"  src: {iframe.get('src','')[:100]}")

# =========================================================
# Step 3: Try downloading one known XLSX and check its content
# =========================================================
print("\n" + "=" * 60)
print("STEP 3: Download & inspect known XLSX from IRDAI")
print("=" * 60)

# Download a life insurance premium XLSX to understand folder structure
test_url = "https://irdai.gov.in/documents/37343/365644/31.01.2025+%E0%A4%95%E0%A5%8B+%E0%A4%9C%E0%A5%80%E0%A4%B5%E0%A4%A8+%E0%A4%AC%E0%A5%80%E0%A4%AE%E0%A4%BE%E0%A4%95%E0%A4%B0%E0%A5%8D%E0%A4%A4%E0%A4%BE%E0%A4%93%E0%A4%82+%E0%A4%95%E0%A4%BE+%E0%A4%AA%E0%A5%8D%E0%A4%B0%E0%A4%A5%E0%A4%AE+%E0%A4%B5%E0%A4%B0%E0%A5%8D%E0%A4%B7+%E0%A4%95%E0%A4%BE+%E0%A4%AA%E0%A5%8D%E0%A4%B0%E0%A5%80%E0%A4%AE%E0%A4%BF%E0%A4%AF%E0%A4%AE+_+First+year+premium+of+Life+Insurers+as+on+31.01.2025.xlsx/39c48931-b067-dcb0-9d8c-1967c0db8429?version=1.0&t=1739782928657&download=true"

r = get(test_url, timeout=20)
if r and r.status_code == 200 and len(r.content) > 5000:
    fname = "docs/fetched/irdai_test_premium.xlsx"
    with open(fname, 'wb') as f:
        f.write(r.content)
    print(f"Downloaded XLSX: {len(r.content)} bytes -> {fname}")
    
    # Try reading with openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fname)
        print(f"Sheets: {wb.sheetnames}")
        ws = wb.active
        print(f"Sheet dimensions: {ws.dimensions}")
        print(f"First few rows:")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 5:
                print(f"  Row {i+1}: {[str(c)[:20] if c else '' for c in row[:8]]}")
    except Exception as e:
        print(f"openpyxl error: {e}")
        # Try pandas
        try:
            import pandas as pd
            df = pd.read_excel(fname)
            print(f"Columns: {list(df.columns)[:10]}")
            print(df.head(3))
        except Exception as e2:
            print(f"pandas error: {e2}")
else:
    print(f"Download failed: {r.status_code if r else 'No response'}")

# =========================================================
# Step 4: Try IRDAI Claims specific page
# =========================================================
print("\n" + "=" * 60)
print("STEP 4: IRDAI Claims page deep scan")
print("=" * 60)

r = get("https://irdai.gov.in/en/claims")
if r and r.status_code == 200:
    soup = BeautifulSoup(r.text, 'html.parser')
    # Get all text content to understand what's there
    text = soup.get_text(separator='\n', strip=True)
    # Find lines with numbers/percentages (likely claims data)
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    claim_lines = [l for l in lines if any(k in l.lower() for k in 
                   ['claim','death','settlement','paid','lic','individual','group']) and len(l) > 10]
    print(f"Claims-related text lines ({len(claim_lines)}):")
    for l in claim_lines[:20]:
        print(f"  {l[:100]}")

print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"Claim links found from folder search: {len(claim_links)}")
for l in claim_links:
    print(f"  [{l['text'][:60]}] -> {l['url'][:80]}")
