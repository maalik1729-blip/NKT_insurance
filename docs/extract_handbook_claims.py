"""
Extract complete claims data from IRDAI Handbook Excel files
Tables 16, 17, 18, 19 = Individual & Group Death Claims with CSR
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, pandas as pd

def find_part1_files():
    files = []
    for root, dirs, fs in os.walk('docs/fetched/handbook'):
        for f in fs:
            fp = os.path.join(root, f)
            fl = f.lower()
            if fl.endswith('.xlsx') and 'hindi' not in fl and 'hin' not in fl.replace('handbook',''):
                if 'part i' in fl or 'part_i' in fl or 'part i' in fp.lower().replace('handbook',''):
                    files.append(fp)
    return sorted(files)

def get_year_from_path(fp):
    for y in ['2024-25','2023-24','2022-23','2021-22','2020-21','2019-20']:
        if y in fp:
            return y
    return 'unknown'

def extract_table_data(ws, target_keywords):
    """Extract data from a worksheet containing claim tables."""
    rows = list(ws.iter_rows(values_only=True))
    result = {}
    
    for i, row in enumerate(rows):
        flat = ' '.join(str(c) for c in row if c).lower()
        
        # Find the main insurer table
        if 'insurer' in flat and ('claim' in flat or 'csr' in flat or 'settlement' in flat):
            # Found header row - extract data below
            header_row = i
            data_rows = rows[header_row+1:header_row+50]
            
            for drow in data_rows:
                if drow is None or all(c is None for c in drow): continue
                a = str(drow[0]).strip() if drow[0] else ''
                if a and a not in ['None',''] and not a.isdigit():
                    insurer = a
                    result[insurer] = drow
            break
    
    return result

def safe_float(v):
    if v is None or str(v).strip() in ['', 'None', '-', 'N/A']: return 0.0
    try: return float(str(v).replace(',',''))
    except: return 0.0

def safe_int(v):
    if v is None or str(v).strip() in ['', 'None', '-', 'N/A']: return 0
    try: return int(float(str(v).replace(',','')))
    except: return 0

# Target sheets known to have claims data
CLAIM_TABLES = {
    'Table 16': 'Individual Death Claims',
    'Table 17': 'Group Death Claims',
    'Table 14': 'Individual Death Claims',   # older format
    'Table 15': 'Group Death Claims',        # older format
}

all_records = []
TARGET_YEARS = ['2019-20','2020-21','2021-22','2022-23','2023-24','2024-25']

part1_files = find_part1_files()
print(f'Found {len(part1_files)} Part I files')

for fpath in part1_files:
    yr_label = get_year_from_path(fpath)
    print(f'\n=== Processing: {yr_label} | {fpath[-55:]} ===')
    
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=5))
            flat_header = ' '.join(str(c) for row in rows for c in row if c).lower()
            
            # Identify claim table type
            is_individual_dc = any(k in flat_header for k in 
                ['individual death', 'indl-dc', 'individual claim', 'table 16', 'table 14'])
            is_group_dc = any(k in flat_header for k in 
                ['group death', 'group dc', 'table 17', 'table 15'])
            is_csr_table = any(k in flat_header for k in 
                ['claim settlement ratio', 'csr', 'settlement ratio'])
            
            if not (is_individual_dc or is_group_dc or is_csr_table):
                continue
            
            category = 'Individual Death Claims' if is_individual_dc else ('Group Death Claims' if is_group_dc else 'CSR Table')
            print(f'\n  Sheet [{sheet_name}] → {category}')
            
            # Get all data rows
            all_rows = list(ws.iter_rows(values_only=True))
            
            # Find the data - look for rows with insurer names and numbers
            current_insurer = None
            current_year = None
            
            for i, row in enumerate(all_rows):
                if not row or all(c is None for c in row): continue
                
                a = str(row[0]).strip() if row[0] is not None else ''
                b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                
                # Update insurer
                if a and a not in ['None','','S.No.','Particulars','Insurer'] and not a.isdigit():
                    try:
                        float(a)
                    except:
                        current_insurer = a
                
                # Check if col B has a year
                for yr in TARGET_YEARS:
                    if yr == b or yr in str(row[0] or ''):
                        current_year = yr
                        break
                
                if not current_insurer or not current_year:
                    continue
                
                # Try to extract numeric data
                cols = list(row)
                while len(cols) < 24: cols.append(None)
                
                # Look for the settlement ratio or paid/total counts
                tn = safe_float(cols[6])  # total
                pn = safe_float(cols[8])  # paid
                
                if tn > 0 and pn > 0:
                    csr = round(pn / tn * 100, 2)
                    rec = {
                        'life_insurer': current_insurer,
                        'year': current_year,
                        'category': category,
                        'source_handbook': yr_label,
                        'total_claims_no': int(tn),
                        'claims_paid_no': int(pn),
                        'claims_intimated_no': safe_int(cols[4]),
                        'claims_pending_start_no': safe_int(cols[2]),
                        'claims_pending_end_no': safe_int(cols[16]),
                        'claims_repudiated_no': safe_int(cols[10]),
                        'claims_rejected_no': safe_int(cols[12]),
                        'claims_paid_ratio_no': csr / 100,
                    }
                    all_records.append(rec)
                    
                    if any(k in current_insurer.upper() for k in ['LIC','TOTAL','INDUSTRY','PRIVATE','PVT']):
                        print(f'    {current_insurer[:20]:20} | {current_year} | Total={int(tn):>10,} | Paid={int(pn):>10,} | CSR={csr:.2f}%')
    
    except Exception as e:
        print(f'  Error: {e}')

print('\n' + '='*65)
print('SUMMARY')
print('='*65)

if all_records:
    df = pd.DataFrame(all_records)
    df = df.drop_duplicates(subset=['life_insurer','year','category'])
    
    print(f'Total records: {len(df)}')
    print(f'Years: {sorted(df.year.unique())}')
    print(f'Categories: {sorted(df.category.unique())}')
    
    # LIC summary
    print('\nLIC Claims Summary:')
    lic = df[df.life_insurer.str.contains('LIC',na=False, case=False)].sort_values(['category','year'])
    for _, r in lic.iterrows():
        print(f"  {r['category']:35} | {r['year']} | Intimated={r['claims_intimated_no']:>12,} | Paid={r['claims_paid_no']:>12,} | CSR={r['claims_paid_ratio_no']*100:.2f}%")
    
    df.to_csv('docs/fetched/handbook_complete_claims.csv', index=False)
    print(f'\nSaved: docs/fetched/handbook_complete_claims.csv')
else:
    print('No records extracted — checking for CSR-only table format...')
    
    # Fallback: scan for CSR table (Table 16 format with ratio columns)
    for fpath in part1_files:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            rows = list(ws.iter_rows(values_only=True, max_row=3))
            flat = ' '.join(str(c) for row in rows for c in row if c).lower()
            if 'claim settlement' in flat or 'csr' in flat:
                print(f'\nCSR Table in [{s}] of {fpath[-40:]}:')
                for row in list(ws.iter_rows(values_only=True))[:10]:
                    clean = [str(c)[:18] if c else '' for c in list(row)[:10]]
                    if any(clean): print(f'  {clean}')
