"""
IRDAI Complete Claims Data Extractor
- Scans ALL claims pages
- Downloads Individual DC, Group DC, Survival Benefits, Maturity Claims
- Covers 2019 to 2024
"""
import sys, io, requests, os, zipfile
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from bs4 import BeautifulSoup
import openpyxl, pandas as pd

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
BASE = "https://irdai.gov.in"
os.makedirs("docs/fetched/all_claims", exist_ok=True)

def get(url, timeout=20):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        return r
    except Exception as e:
        print(f"  ERR [{url[-40:]}]: {e}")
        return None

def find_all_file_links(html):
    soup = BeautifulSoup(html, 'html.parser')
    links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        txt = a.get_text(strip=True)
        if any(x in href.lower() for x in ['.xlsx','.xls','.zip','.csv','.pdf']):
            full = href if href.startswith('http') else BASE + href
            links.append({'text': txt, 'url': full})
    return links

# ================================================================
# STEP 1: Comprehensive IRDAI pages scan
# ================================================================
print("=" * 70)
print("STEP 1: Scanning ALL IRDAI Claims Pages")
print("=" * 70)

all_claim_pages = [
    "https://irdai.gov.in/en/claims",
    "https://irdai.gov.in/en/life",
    "https://irdai.gov.in/en/life1",
    "https://irdai.gov.in/en/life3",
    "https://irdai.gov.in/en/life-reports",
    "https://irdai.gov.in/en/non-life",
    "https://irdai.gov.in/en/non-life1",
    "https://irdai.gov.in/en/non-life2",
    "https://irdai.gov.in/en/non-life-reports",
    "https://irdai.gov.in/en/report-and-statistics1",
    "https://irdai.gov.in/en/annual-accounts",
    "https://irdai.gov.in/en/handbooks",
    "https://irdai.gov.in/en/finance-accounts-life-",
    "https://irdai.gov.in/en/finance-accounts-non-life-",
]

all_downloads = {}  # url -> {text, page}

for page_url in all_claim_pages:
    r = get(page_url)
    if not r or r.status_code != 200: continue
    links = find_all_file_links(r.text)
    claim_links = [l for l in links if any(k in (l['text']+l['url']).lower()
                   for k in ['claim','death','survival','maturity','benefit','settlement',
                             '2019','2020','2021','2022','2023','2024','2025','2026',
                             'life','individual','group'])]
    if claim_links:
        print(f"\nPage: {page_url[-40:]}")
        for l in claim_links[:15]:
            if l['url'] not in all_downloads:
                all_downloads[l['url']] = {'text': l['text'], 'page': page_url}
                print(f"  [{l['text'][:60]}]")
                print(f"   -> {l['url'][:100]}")

# ================================================================
# STEP 2: Check IRDAI document folder 501918 (known claims folder)
# Also try nearby folders
# ================================================================
print("\n" + "=" * 70)
print("STEP 2: Scanning IRDAI Document Folders for Claims")
print("=" * 70)

# Known folder for death claims: 501918
# Try this folder and nearby ones
for folder_id in [501918, 501919, 501920, 501950, 502000, 503000, 504000, 505000]:
    url = f"{BASE}/documents/37343/{folder_id}"
    r = get(url, timeout=10)
    if r and r.status_code == 200 and len(r.content) > 20000:
        links = find_all_file_links(r.text)
        if links:
            print(f"\nFolder {folder_id}:")
            for l in links[:10]:
                if l['url'] not in all_downloads:
                    all_downloads[l['url']] = {'text': l['text'], 'page': f"folder/{folder_id}"}
                    print(f"  [{l['text'][:60]}] -> {l['url'][:90]}")

# ================================================================
# STEP 3: Download ALL found XLSX files
# ================================================================
print("\n" + "=" * 70)
print("STEP 3: Downloading All Claims Files")
print("=" * 70)

xlsx_files = {url: info for url, info in all_downloads.items()
              if any(x in url.lower() for x in ['.xlsx', '.xls', '.zip'])}

print(f"Total XLSX/ZIP files to download: {len(xlsx_files)}")

downloaded_files = {}
for url, info in xlsx_files.items():
    txt = info['text']
    # Only download relevant files
    if not any(k in (txt+url).lower() for k in
               ['claim','death','survival','maturity','benefit','settlement','life','individual']):
        continue
    
    safe_name = txt[:40].replace('/', '_').replace(' ', '_').replace('\\', '_').strip('_')
    if not safe_name:
        safe_name = url.split('/')[-1][:40]
    
    fname = f"docs/fetched/all_claims/{safe_name}.xlsx"
    if os.path.exists(fname):
        print(f"  Already exists: {fname}")
        downloaded_files[fname] = txt
        continue
    
    print(f"  Downloading: {txt[:50]}...")
    r = get(url, timeout=30)
    if r and r.status_code == 200 and len(r.content) > 5000:
        ctype = r.headers.get('Content-Type', '')
        if 'html' not in ctype.lower():
            with open(fname, 'wb') as f:
                f.write(r.content)
            print(f"  Saved: {fname} ({len(r.content):,} bytes)")
            downloaded_files[fname] = txt
        else:
            print(f"  HTML response (JS-rendered) - skipped")
    else:
        print(f"  Failed: {r.status_code if r else 'no response'}")

# ================================================================
# STEP 4: Extract ALL data from downloaded files
# ================================================================
print("\n" + "=" * 70)
print("STEP 4: Extracting Data from Downloaded Files")
print("=" * 70)

all_data = []
TARGET_YEARS = ['2019-20', '2020-21', '2021-22', '2022-23', '2023-24', '2024-25']

def extract_claims_from_sheet(ws, year_filter=None):
    """Generic extractor for any claims sheet."""
    rows_out = []
    insurer = None
    all_ws_rows = list(ws.iter_rows(values_only=True))
    
    for row in all_ws_rows:
        if not row or all(c is None for c in row): continue
        a = str(row[0]).strip() if row[0] is not None else ''
        b = str(row[1]).strip() if row[1] is not None else ''
        
        # Update insurer name
        if a and a not in ['None', 'Life Insurer', ''] and not a[0].isdigit():
            try: float(a)
            except: insurer = a
        
        # Filter by year
        if year_filter and b != year_filter: continue
        if not b or '-' not in b or not insurer: continue
        
        def f(v):
            try: return float(v) if v is not None else 0.0
            except: return 0.0
        
        c = list(row)
        while len(c) < 24: c.append(None)
        tn = f(c[6]); pn = f(c[8]); ta = f(c[7]); pa = f(c[9])
        if tn == 0: continue
        
        rows_out.append({
            'life_insurer': insurer,
            'year': b,
            'claims_pending_start_no': int(f(c[2])),
            'claims_pending_start_amt': f(c[3]),
            'claims_intimated_no': int(f(c[4])),
            'claims_intimated_amt': f(c[5]),
            'total_claims_no': int(tn),
            'total_claims_amt': ta,
            'claims_paid_no': int(pn),
            'claims_paid_amt': pa,
            'claims_repudiated_no': int(f(c[10])),
            'claims_repudiated_amt': f(c[11]),
            'claims_rejected_no': int(f(c[12])),
            'claims_rejected_amt': f(c[13]),
            'claims_pending_end_no': int(f(c[16])),
            'claims_pending_end_amt': f(c[17]),
            'claims_paid_ratio_no': round(pn/tn, 4),
            'claims_paid_ratio_amt': round(pa/ta, 4) if ta > 0 else 0.0,
        })
    return rows_out

# Process all downloaded XLSX files
for fpath, title in list(downloaded_files.items()) + [
    ('docs/fetched/death_claims_2021-22.xlsx', 'Death Claims 2021-22'),
    ('docs/fetched/death_claims_2020-21.xlsx', 'Death Claims 2020-21'),
    ('docs/fetched/death_claims_2019-20.xlsx', 'Death Claims 2019-20'),
    ('docs/fetched/death_claims_2018-19.xlsx', 'Death Claims 2018-19'),
]:
    if not os.path.exists(fpath): continue
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"\nFile: {fpath[-50:]} | Sheets: {wb.sheetnames}")
        
        sheet_category_map = {
            'Indl-DC': 'Individual Death Claims',
            'Group DC': 'Group Death Claims',
            'Indl-SB': 'Individual Survival Benefit Claims',
            'Indl-MC': 'Individual Maturity Claims',
            'Non-Life': 'Non-Life Claims',
        }
        
        for sheet_name in wb.sheetnames:
            category = sheet_category_map.get(sheet_name, sheet_name)
            ws = wb[sheet_name]
            
            # Get all years in this sheet
            years_in_sheet = set()
            for row in ws.iter_rows(values_only=True):
                if row[1] and str(row[1]).strip() not in ['None','Year','']:
                    years_in_sheet.add(str(row[1]).strip())
            
            for yr in years_in_sheet:
                if yr not in TARGET_YEARS: continue
                rows = extract_claims_from_sheet(ws, yr)
                for r in rows:
                    r['category'] = category
                    r['source_file'] = fpath.split('/')[-1]
                all_data.extend(rows)
                print(f"  Sheet '{sheet_name}' | Year {yr}: {len(rows)} rows")
    except Exception as e:
        print(f"  Error: {e}")

# ================================================================
# STEP 5: Save final dataset
# ================================================================
print("\n" + "=" * 70)
print("STEP 5: Saving Final Dataset")
print("=" * 70)

if all_data:
    df = pd.DataFrame(all_data)
    df = df.drop_duplicates(subset=['life_insurer', 'year', 'category'])
    df = df.sort_values(['category', 'year', 'life_insurer'])
    
    print(f"Total rows: {len(df)}")
    print(f"Years: {sorted(df.year.unique())}")
    print(f"Categories: {sorted(df.category.unique())}")
    
    df.to_csv('docs/fetched/complete_irdai_claims_2019_2024.csv', index=False)
    print(f"\nSaved: docs/fetched/complete_irdai_claims_2019_2024.csv")
    
    # Summary by category and year
    print("\nSummary:")
    summary = df.groupby(['category', 'year'])['claims_intimated_no'].sum().reset_index()
    print(summary.to_string(index=False))
    
    # LIC specific
    print("\nLIC Data:")
    lic = df[df.life_insurer.str.contains('LIC', na=False)][
        ['category', 'year', 'claims_intimated_no', 'claims_paid_no', 'claims_paid_ratio_no']
    ].sort_values(['category', 'year'])
    print(lic.to_string(index=False))
else:
    print("No data collected!")

print("\n" + "=" * 70)
print("COMPLETE!")
print("=" * 70)
