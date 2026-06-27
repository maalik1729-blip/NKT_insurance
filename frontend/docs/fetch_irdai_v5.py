"""
IRDAI Death Claims Downloader & CSV Converter
Downloads all available death claims XLSX from IRDAI
and converts to same format as existing dataset.
"""
import sys, io, requests, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import openpyxl

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
os.makedirs("docs/fetched", exist_ok=True)

# =========================================================
# Known XLSX download URLs from IRDAI claims page
# =========================================================
DEATH_CLAIM_FILES = {
    "2021-22": "https://irdai.gov.in/documents/37343/501918/Death+Claims_upto++March+2022_website.xlsx/275cef49-f4ea-3756-9e5b-ac9d30164a60?version=1.2&t=1664971947224&download=true",
    "2020-21": "https://irdai.gov.in/documents/37343/501918/Death+Claim_2020-21.xlsx/fedc5295-8776-d212-4123-575abff9c4ca?version=1.26&t=1664971912807&download=true",
    "2019-20": "https://irdai.gov.in/documents/37343/501918/Death+Claim_2019-20.xlsx/d3c07c8a-4a05-1e1a-7ae0-7102fac863c0?version=1.2&t=1664971931029&download=true",
    "2018-19": "https://irdai.gov.in/documents/37343/501918/Death+Claim_2018-19.xlsx/d80d2da1-ffe9-89ce-91d9-2de1f626cd60?version=1.2&t=1664971941085&download=true",
}

# Also check page 2 for 2022-23 and 2023-24
PAGE2_URL = "https://irdai.gov.in/en/claims?p_p_id=com_irdai_document_media_IRDAIDocumentMediaPortlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&_com_irdai_document_media_IRDAIDocumentMediaPortlet_cur=2&_com_irdai_document_media_IRDAIDocumentMediaPortlet_delta=20&_com_irdai_document_media_IRDAIDocumentMediaPortlet_orderByCol=dateid_String_sortable&_com_irdai_document_media_IRDAIDocumentMediaPortlet_orderByType=desc"

print("=" * 60)
print("Checking page 2 for newer death claims data...")
print("=" * 60)
try:
    from bs4 import BeautifulSoup
    r = requests.get(PAGE2_URL, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, 'html.parser')
    for a in soup.find_all('a', href=True):
        href = a['href']
        txt = a.get_text(strip=True)
        if txt and any(x in (txt+href).lower() for x in ['death','claim','xlsx','2022','2023','2024']):
            full = href if href.startswith('http') else 'https://irdai.gov.in' + href
            print(f"[{txt[:70]}] -> {full[:100]}")
            if '.xlsx' in href.lower() and ('2022' in txt or '2023' in txt or '2024' in txt):
                year = '2022-23' if '2022' in txt or '2023' in txt else '2023-24'
                DEATH_CLAIM_FILES[year] = full
except Exception as e:
    print(f"Page 2 check error: {e}")

# Also try page sorted by newest first
SORTED_URL = "https://irdai.gov.in/en/claims?p_p_id=com_irdai_document_media_IRDAIDocumentMediaPortlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&_com_irdai_document_media_IRDAIDocumentMediaPortlet_cur=1&_com_irdai_document_media_IRDAIDocumentMediaPortlet_delta=20&_com_irdai_document_media_IRDAIDocumentMediaPortlet_orderByCol=dateid_String_sortable&_com_irdai_document_media_IRDAIDocumentMediaPortlet_orderByType=desc"
print("\nChecking sorted by date (newest first)...")
try:
    r = requests.get(SORTED_URL, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, 'html.parser')
    for a in soup.find_all('a', href=True):
        href = a['href']
        txt = a.get_text(strip=True)
        if txt and any(x in (txt+href).lower() for x in ['death','claim','xlsx']):
            full = href if href.startswith('http') else 'https://irdai.gov.in' + href
            print(f"[{txt[:70]}] -> {full[:110]}")
            if '.xlsx' in href.lower():
                for yr in ['2022-23','2023-24','2024-25']:
                    if yr[:4] in txt or yr[5:] in txt:
                        DEATH_CLAIM_FILES[yr] = full
                        print(f"  >>> Added as {yr}")
except Exception as e:
    print(f"Sorted URL error: {e}")

# =========================================================
# Download all XLSX files
# =========================================================
print("\n" + "=" * 60)
print("Downloading IRDAI Death Claims XLSX files...")
print("=" * 60)

downloaded = {}
for year, url in DEATH_CLAIM_FILES.items():
    print(f"\nDownloading {year}...")
    try:
        r = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True)
        if r.status_code == 200 and len(r.content) > 5000:
            fname = f"docs/fetched/death_claims_{year}.xlsx"
            with open(fname, 'wb') as f:
                f.write(r.content)
            print(f"  Saved: {fname} ({len(r.content):,} bytes)")
            downloaded[year] = fname
        else:
            print(f"  Failed: status={r.status_code}, size={len(r.content)}")
    except Exception as e:
        print(f"  Error: {e}")

# =========================================================
# Inspect downloaded files
# =========================================================
print("\n" + "=" * 60)
print("Inspecting downloaded files...")
print("=" * 60)

for year, fname in downloaded.items():
    try:
        wb = openpyxl.load_workbook(fname, data_only=True)
        print(f"\nYear: {year} | File: {fname}")
        print(f"  Sheets: {wb.sheetnames}")
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            print(f"\n  Sheet: '{sheet_name}' | Dims: {ws.dimensions}")
            print(f"  First 6 rows:")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 6:
                    row_data = [str(c)[:18] if c is not None else '' for c in list(row)[:10]]
                    if any(row_data):
                        print(f"    {i+1}: {row_data}")
    except Exception as e:
        print(f"  Error reading {year}: {e}")

print("\n" + "=" * 60)
print(f"Total files downloaded: {len(downloaded)}")
for y, f in downloaded.items():
    print(f"  {y}: {f}")
print("=" * 60)
