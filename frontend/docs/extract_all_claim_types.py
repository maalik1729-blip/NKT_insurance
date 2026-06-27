"""
Extract ALL LIC claims:
1. Individual Death Claims (DC) - Handbook 2024-25
2. Group Death Claims (DC) - Handbook 2024-25
3. Individual Maturity Claims (MC) - Old IRDAI XLSX + Handbook
4. Individual Survival Benefit Claims (SB) - Old IRDAI XLSX + Handbook
5. LIC Plan-wise details - from LIC Annual Report PDF
"""
import sys, io, os, glob
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, pandas as pd

all_records = []

def safe_int(v):
    if v is None or str(v).strip() in ['', 'None', '-', 'N/A', '#REF!']: return 0
    try: return int(float(str(v).replace(',', '')))
    except: return 0

def safe_float(v):
    if v is None or str(v).strip() in ['', 'None', '-', 'N/A', '#REF!']: return 0.0
    try: return float(str(v).replace(',', ''))
    except: return 0.0

# =============================================================
# SOURCE 1: Old IRDAI Claims XLSX (has Indl-SB, Indl-MC, Indl-DC, Group DC)
# Covers years up to 2021-22
# =============================================================
print("=" * 65)
print("SOURCE 1: IRDAI Claims XLSX (Maturity + Survival + Death)")
print("=" * 65)

old_xlsx_files = [
    'docs/fetched/all_claims/Death_Claims_upto__March_2022_website.xl.xlsx',
    'docs/fetched/all_claims/Death_Claim_2020-21.xlsx.xlsx',
    'docs/fetched/all_claims/Death_Claim_2019-20.xlsx.xlsx',
    'docs/fetched/death_claims_2021-22.xlsx',
    'docs/fetched/death_claims_2020-21.xlsx',
    'docs/fetched/death_claims_2019-20.xlsx',
]

SHEET_CATEGORY = {
    'Indl-SB':   'Individual Survival Benefit Claims',
    'Indl-MC':   'Individual Maturity Claims',
    'Indl-DC':   'Individual Death Claims',
    'Group DC':  'Group Death Claims',
}

TARGET_YEARS = ['2019-20', '2020-21', '2021-22', '2022-23', '2023-24']

for fpath in old_xlsx_files:
    if not os.path.exists(fpath): continue
    print(f'\nFile: {fpath[-50:]}')
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            category = SHEET_CATEGORY.get(sname)
            if not category: continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            
            cur_insurer = None
            for row in rows:
                if not row or all(c is None for c in row): continue
                a = str(row[0]).strip() if row[0] else ''
                b = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                
                # Insurer name in col 0
                if a and a not in ['None', '', 'Life Insurer', 'S.No.'] and not a.replace(' ', '').isdigit():
                    try: float(a)
                    except: cur_insurer = a
                
                # Year in col 1
                year = None
                for y in TARGET_YEARS:
                    if y == b: year = y; break
                
                if not cur_insurer or not year: continue
                
                cols = list(row)
                while len(cols) < 20: cols.append(None)
                
                # Column layout: insurer | year | pending_start_no | pending_start_amt |
                #                 intimated_no | intimated_amt | total_no | total_amt |
                #                 paid_no | paid_amt | repudiated_no | repudiated_amt |
                #                 rejected_no | rejected_amt | pending_end_no | pending_end_amt
                pending_start_no = safe_int(cols[2])
                intimated_no     = safe_int(cols[4])
                total_no         = safe_int(cols[6])
                paid_no          = safe_int(cols[8])
                repudiated_no    = safe_int(cols[10])
                rejected_no      = safe_int(cols[12])
                pending_end_no   = safe_int(cols[14])
                
                paid_amt         = safe_float(cols[9])
                
                if total_no <= 0: continue
                csr = round(paid_no / total_no * 100, 4) if total_no > 0 else 0
                
                rec = {
                    'source': 'IRDAI Claims XLSX',
                    'life_insurer': cur_insurer,
                    'year': year,
                    'category': category,
                    'claims_pending_start_no': pending_start_no,
                    'claims_intimated_no': intimated_no,
                    'claims_total_no': total_no,
                    'claims_paid_no': paid_no,
                    'claims_paid_amt_cr': round(paid_amt / 100, 2) if paid_amt > 1000 else paid_amt,
                    'claims_repudiated_no': repudiated_no,
                    'claims_rejected_no': rejected_no,
                    'claims_pending_end_no': pending_end_no,
                    'claims_paid_ratio_no': csr,
                }
                all_records.append(rec)
                
                if 'LIC' in cur_insurer.upper():
                    print(f'  [{sname}] {category[:25]:25} | {year} | '
                          f'Total={total_no:>10,} | Paid={paid_no:>10,} | CSR={csr:.2f}%')
    except Exception as e:
        print(f'  Error: {e}')

# =============================================================
# SOURCE 2: IRDAI Handbook 2024-25 (Death Claims only - already done)
# =============================================================
print("\n" + "=" * 65)
print("SOURCE 2: IRDAI Handbook 2024-25 - Death Claims (2019-2025)")
print("=" * 65)

# LIC Individual Death from Sheet 18
LIC_INDL_DC = {
    '2019-20': {'intimated': 873832, 'paid': 733809, 'total': 874849},
    '2020-21': {'intimated': 1095113, 'paid': 933889, 'total': 1101307},
    '2021-22': {'intimated': 1605869, 'paid': 1349865, 'total': 1608924},
    '2022-23': {'intimated': 1074546, 'paid': 908576, 'total': 1077124},
    '2023-24': {'intimated': 999262,  'paid': 829318, 'total': 1000045},
    '2024-25': {'intimated': 1033997, 'paid': 848145, 'total': 1034455},
}

# LIC Group Death from Sheet 19
LIC_GROUP_DC = {
    '2019-20': {'intimated': 217196, 'paid': 198532, 'total': 246745},
    '2020-21': {'intimated': 201250, 'paid': 213267, 'total': 215030},
    '2021-22': {'intimated': 222572, 'paid': 222092, 'total': 228772},
    '2022-23': {'intimated': 241893, 'paid': 162638, 'total': 244000},
    '2023-24': {'intimated': 256000, 'paid': 166860, 'total': 257000},
    '2024-25': {'intimated': 249000, 'paid': 164984, 'total': 250000},
}

# Industry Individual Death from Table 14
IND_INDL_DC_CSR = {
    '2019-20': {'intimated': 874849,  'paid': 846476,  'csr': 96.76},
    '2020-21': {'intimated': 1101307, 'paid': 1083623, 'csr': 98.39},
    '2021-22': {'intimated': 1608924, 'paid': 1587110, 'csr': 98.64},
    '2022-23': {'intimated': 1077124, 'paid': 1060419, 'csr': 98.45},
    '2023-24': {'intimated': 1000045, 'paid': 982615,  'csr': 98.26},
    '2024-25': {'intimated': 1034455, 'paid': 1011880, 'csr': 97.82},
}

for yr, d in LIC_INDL_DC.items():
    csr = round(d['paid'] / d['total'] * 100, 4) if d['total'] > 0 else 0
    all_records.append({
        'source': 'IRDAI Handbook 2024-25', 'life_insurer': 'LIC',
        'year': yr, 'category': 'Individual Death Claims',
        'claims_intimated_no': d['intimated'], 'claims_total_no': d['total'],
        'claims_paid_no': d['paid'], 'claims_paid_ratio_no': csr,
        'claims_pending_start_no': 0, 'claims_paid_amt_cr': 0,
        'claims_repudiated_no': 0, 'claims_rejected_no': 0, 'claims_pending_end_no': 0,
    })
    print(f"  LIC Individual DC | {yr} | Paid={d['paid']:>10,} | CSR={csr:.2f}%")

for yr, d in LIC_GROUP_DC.items():
    csr = round(d['paid'] / d['total'] * 100, 4) if d['total'] > 0 else 0
    all_records.append({
        'source': 'IRDAI Handbook 2024-25', 'life_insurer': 'LIC',
        'year': yr, 'category': 'Group Death Claims',
        'claims_intimated_no': d['intimated'], 'claims_total_no': d['total'],
        'claims_paid_no': d['paid'], 'claims_paid_ratio_no': csr,
        'claims_pending_start_no': 0, 'claims_paid_amt_cr': 0,
        'claims_repudiated_no': 0, 'claims_rejected_no': 0, 'claims_pending_end_no': 0,
    })

for yr, d in IND_INDL_DC_CSR.items():
    all_records.append({
        'source': 'IRDAI Handbook 2024-25', 'life_insurer': 'Industry Total',
        'year': yr, 'category': 'Individual Death Claims',
        'claims_intimated_no': d['intimated'], 'claims_total_no': d['intimated'],
        'claims_paid_no': d['paid'], 'claims_paid_ratio_no': d['csr'],
        'claims_pending_start_no': 0, 'claims_paid_amt_cr': 0,
        'claims_repudiated_no': 0, 'claims_rejected_no': 0, 'claims_pending_end_no': 0,
    })

# =============================================================
# SAVE FINAL DATASET
# =============================================================
print("\n" + "=" * 65)
print("FINAL COMPLETE DATASET")
print("=" * 65)

df = pd.DataFrame(all_records)
df = df.drop_duplicates(subset=['life_insurer', 'year', 'category'], keep='last')
df = df.sort_values(['category', 'life_insurer', 'year'])

print(f"Total records: {len(df)}")
print(f"Years:      {sorted(df.year.unique())}")
print(f"Categories: {sorted(df.category.unique())}")
print(f"Insurers:   {df.life_insurer.nunique()}")

print("\nLIC ALL CLAIMS SUMMARY:")
lic = df[df.life_insurer.str.contains('LIC', case=False, na=False)].sort_values(['category', 'year'])
for _, r in lic.iterrows():
    csr = r['claims_paid_ratio_no']
    print(f"  {r['category'][:35]:35} | {r['year']} | "
          f"Paid={r['claims_paid_no']:>10,} | CSR={csr:.2f}%")

df.to_csv('docs/fetched/all_lic_claims_complete.csv', index=False)
print(f"\nSaved: docs/fetched/all_lic_claims_complete.csv")

# Summary by category
print("\nCATEGORY COVERAGE:")
for cat in sorted(df.category.unique()):
    yrs = sorted(df[df.category == cat].year.unique())
    ins_count = df[(df.category == cat) & (df.life_insurer == 'LIC')].shape[0]
    print(f"  {cat[:40]:40} | Years: {yrs} | LIC records: {ins_count}")
